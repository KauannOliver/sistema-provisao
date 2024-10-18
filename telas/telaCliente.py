import flet as ft
from formularios.cadCliente import criar_formulario_cliente
import pandas as pd
from openpyxl import load_workbook

# Função para carregar os dados da sheet 'Impostos'
def carregar_dados_cliente(caminho_arquivo, sheet_name='Impostos'):
    try:
        df = pd.read_excel(caminho_arquivo, sheet_name=sheet_name)
        return df
    except Exception as e:
        print(f"Erro ao carregar dados: {e}")
        return pd.DataFrame()

# Função para converter os valores em porcentagem
def formatar_porcentagem(valor):
    try:
        return f"{float(valor) * 100:.2f}%"
    except ValueError:
        return "0.00%"

# Função para mostrar a tabela de clientes com paginação
def TelaCliente(page, container):
    cadastros_container = ft.Container(padding=20)
    itens_por_pagina = 10
    pagina_atual = 1
    total_paginas = 1

    # Carregar os dados da sheet 'Impostos'
    caminho_arquivo = "banco/ProvisaoBD.xlsx"  # Ajuste o caminho conforme necessário
    global df  # Definir df como uma variável global
    df = carregar_dados_cliente(caminho_arquivo)

    total_registros = len(df)
    total_paginas = (total_registros + itens_por_pagina - 1) // itens_por_pagina

    def handle_page_change(e):
        nonlocal pagina_atual
        if e.control.data == "prev" and pagina_atual > 1:
            pagina_atual -= 1
        elif e.control.data == "next" and pagina_atual < total_paginas:
            pagina_atual += 1
        mostrar_cadastros()

    def mostrar_cadastros():
        nonlocal total_paginas
        inicio = (pagina_atual - 1) * itens_por_pagina
        fim = inicio + itens_por_pagina
        dados_pagina = df.iloc[inicio:fim]

        tabela_dados = []
        for _, row in dados_pagina.iterrows():
            tabela_dados.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Container(
                            content=ft.Text(str(row['CLIENTE']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=230
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(row['UND NEGÓCIO'], text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=160
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(formatar_porcentagem(row['ICMS']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=90
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(formatar_porcentagem(row['ISS']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=90
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(formatar_porcentagem(row['PIS']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=90
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(formatar_porcentagem(row['COFINS']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=90
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(formatar_porcentagem(row['CPRB']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=90
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.DELETE,
                                icon_color="#212A57",
                                on_click=lambda e, row=row: deletar_cliente(row)
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                    ]
                )
            )

        tabela = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Container(
                    content=ft.Text("CLIENTE", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=230
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("UND NEGÓCIO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=160
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("ICMS", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=90
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("ISS", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=90
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("PIS", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=90
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("COFINS", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=90
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("CPRB", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=90
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("AÇÃO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
            ],
            rows=tabela_dados,
            heading_row_color=ft.colors.with_opacity(1, "#f36229"),
            heading_text_style=ft.TextStyle(color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
            border=ft.BorderSide(color=ft.colors.BLACK, width=1),
            width=1000,
            column_spacing=0,
            show_bottom_border=False,
            show_checkbox_column=False,
        )

        paginacao_controls = ft.Row(
            controls=[
                ft.IconButton(
                    icon=ft.icons.ARROW_BACK_IOS,
                    icon_color="black",
                    data="prev",
                    on_click=handle_page_change,
                    disabled=pagina_atual == 1
                ),
                ft.Text(f"{pagina_atual} de {total_paginas}"),
                ft.IconButton(
                    icon=ft.icons.ARROW_FORWARD_IOS,
                    icon_color="black",
                    data="next",
                    on_click=handle_page_change,
                    disabled=pagina_atual == total_paginas
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=20
        )

        cadastros_container.content = ft.Column(
            [tabela, paginacao_controls],
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=20
        )

        page.update()

    # Função para deletar um cliente
    def deletar_cliente(row):
        try:
            # Carregar o arquivo Excel
            caminho_arquivo = "banco/ProvisaoBD.xlsx"
            wb = load_workbook(caminho_arquivo)
            sheet = wb["Impostos"]

            # Procurar a linha que contém o cliente correspondente
            cliente_para_deletar = str(row['CLIENTE']).strip()
            linha_para_deletar = None

            for i, excel_row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cliente_excel = str(excel_row[0]).strip() if excel_row[0] is not None else None
                if cliente_excel == cliente_para_deletar:
                    linha_para_deletar = i
                    break

            if linha_para_deletar:
                # Deletar a linha encontrada
                sheet.delete_rows(linha_para_deletar)
                wb.save(caminho_arquivo)

                # Exibir uma mensagem de sucesso
                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Cliente {cliente_para_deletar} deletado com sucesso!"), bgcolor=ft.colors.GREEN))

                # Atualizar a exibição da tabela
                global df  # Atualizar o DataFrame global após a exclusão
                df = carregar_dados_cliente(caminho_arquivo)  # Recarregar os dados após a exclusão
                mostrar_cadastros()  # Atualizar a tabela
            else:
                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro: Cliente {cliente_para_deletar} não encontrado!"), bgcolor=ft.colors.RED))
        except Exception as e:
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro ao deletar o cliente: {str(e)}"), bgcolor=ft.colors.RED))

    # Função para abrir o modal de cadastro de cliente
    def abrir_cadastro(e):
        # Criar o conteúdo do modal usando o formulário de cadastro
        dialog_content = criar_formulario_cliente(page)

        # Criar o modal
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cadastro de Cliente", weight="bold"),
            content=dialog_content
        )

        # Abrir o modal
        page.dialog = dialog
        dialog.open = True
        page.update()

    # Botão "Cadastrar" na parte superior direita
    botoes_superiores = ft.Row(
        controls=[
            ft.ElevatedButton(
                text="Cadastrar",
                on_click=abrir_cadastro,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            )
        ],
        width=950,
        alignment=ft.MainAxisAlignment.END,
        spacing=10,
    )

    container.content = ft.Column(
        controls=[
            botoes_superiores,
            cadastros_container,
        ],
        alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        spacing=20,
        width="100%"
    )

    mostrar_cadastros()
    page.update()
