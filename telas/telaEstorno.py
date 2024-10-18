import flet as ft
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from banco.banco import *
from formularios.cadEstorno import criar_formulario_estorno

### função para carregar os dados de estorno do Excel ###
def carregar_dados_estorno(caminho_arquivo, sheet_name='Estornos'):
    try:
        df = pd.read_excel(caminho_arquivo, sheet_name=sheet_name)
        return df
    except Exception as e:
        print(f"Erro ao carregar dados: {e}")
        return pd.DataFrame()

### função para determinar a cor do ícone baseado no valor estornado ###
def determinar_cor_estorno(row):
    ### pegar chave da provisão e garantir que seja string ###
    chave_atual = str(int(row['CHAVE'])) if isinstance(row['CHAVE'], float) else str(row['CHAVE'])
    
    ### carregar dados de estorno e provisões do Excel ###
    estornos_df = pd.read_excel("banco/ProvisaoBD.xlsx", sheet_name='Estornos')
    provisoes_df = pd.read_excel("banco/ProvisaoBD.xlsx", sheet_name='Provisões')

    ### converter chave para string para garantir compatibilidade ###
    estornos_df['CHAVE'] = estornos_df['CHAVE'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))
    provisoes_df['CHAVE'] = provisoes_df['CHAVE'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))

    ### filtrar estornos e provisões com base na chave ###
    estornos_filtrados = estornos_df[estornos_df['CHAVE'] == chave_atual]
    provisao = provisoes_df[provisoes_df['CHAVE'] == chave_atual]

    ### somar o total estornado ###
    total_estornado = estornos_filtrados['VALOR ESTORNADO'].sum()

    ### obter receita bruta da provisão correspondente ###
    receita_bruta = provisao['RECEITA BRUTA'].iloc[0] if not provisao.empty else 0
    if isinstance(receita_bruta, str):
        receita_bruta = receita_bruta.replace('.', '').replace(',', '.')
        receita_bruta = float(receita_bruta)

    ### calcular o valor que falta estornar ###
    faltar_estornar = receita_bruta - total_estornado

    ### definir cor com base no quanto foi estornado ###
    if total_estornado == 0:
        return ft.colors.RED  ### nada estornado ###
    elif faltar_estornar > 0:
        return ft.colors.AMBER  ### estorno parcial ###
    else:
        return ft.colors.GREEN  ### estorno total ###

### função principal da tela de estorno ###
def TelaEstorno(page, container):
    cadastros_container = ft.Container(padding=20)  ### contêiner para mostrar os dados ###
    itens_por_pagina = 10  ### qtde de itens por página ###
    pagina_atual = 1  ### inicializa a página atual ###
    mes_selecionado = None  ### mês inicial sem seleção ###
    ano_selecionado = None  ### ano inicial sem seleção ###
    total_paginas = 1  ### total de páginas inicial como 1 ###

    ### lista de meses para dropdown ###
    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]

    ### lista de anos para dropdown ###
    anos = [str(ano) for ano in range(2023, 2027)]

    ### inicializa o FilePicker ###
    file_picker = ft.FilePicker(on_result=None)
    page.overlay.append(file_picker)

    ### função para manipular a mudança de mês ###
    def handle_mes_change(e):
        nonlocal mes_selecionado, pagina_atual
        mes_selecionado = meses.index(e.control.value) + 1 if e.control.value else None
        pagina_atual = 1
        mostrar_cadastros()

    ### função para manipular a mudança de ano ###
    def handle_ano_change(e):
        nonlocal ano_selecionado, pagina_atual
        ano_selecionado = int(e.control.value) if e.control.value else None
        pagina_atual = 1
        mostrar_cadastros()

    ### função para manipular a mudança de página ###
    def handle_page_change(e):
        nonlocal pagina_atual
        if e.control.data == "prev" and pagina_atual > 1:
            pagina_atual -= 1
        elif e.control.data == "next" and pagina_atual < total_paginas:
            pagina_atual += 1
        mostrar_cadastros()

    ### função para formatar valores como moeda ###
    def format_currency(value):
        return f"R$ {value:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")

    ### função para mostrar os cadastros filtrados e paginados ###
    def mostrar_cadastros():
        nonlocal total_paginas

        ### caminho para o arquivo Excel ###
        caminho_arquivo = "banco/ProvisaoBD.xlsx"
        df = carregar_dados_estorno(caminho_arquivo)

        ### filtrar por mês e ano, se selecionados ###
        if mes_selecionado and ano_selecionado:
            df['DATA ESTORNO'] = pd.to_datetime(df['DATA ESTORNO'], errors='coerce')
            df = df[(df['DATA ESTORNO'].dt.month == mes_selecionado) & (df['DATA ESTORNO'].dt.year == ano_selecionado)]

        total_registros = len(df)
        total_paginas = (total_registros + itens_por_pagina - 1) // itens_por_pagina

        inicio = (pagina_atual - 1) * itens_por_pagina
        fim = inicio + itens_por_pagina
        dados_pagina = df.iloc[inicio:fim]

        tabela_dados = []
        ### loop para adicionar dados à tabela ###
        for _, row in dados_pagina.iterrows():
            tabela_dados.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Container(
                            content=ft.Text(str(row['CHAVE']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=160
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(row['DATA ESTORNO'].strftime('%d/%m/%Y'), text_align=ft.TextAlign.CENTER) if not pd.isnull(row['DATA ESTORNO']) else ft.Text("Data inválida", text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=180
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(format_currency(row['VALOR ESTORNADO']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=200
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.CIRCLE,  ### ícone circle no lugar de editar ###
                                icon_color=determinar_cor_estorno(row),  ### definir cor baseado no valor estornado ###
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.DELETE,
                                icon_color="#212a57",
                                on_click=lambda e, row=row: deletar_estorno(row)
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.VISIBILITY,
                                icon_color="#212a57",
                                on_click=lambda e, row=row: visualizar_estorno(row)
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                    ]
                )
            )

        ### tabela de dados com colunas ###
        tabela = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Container(
                    content=ft.Text("CHAVE", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=160
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("DATA ESTORNO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=180
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("VALOR ESTORNADO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=200
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("AÇÕES", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
            ],
            rows=tabela_dados,
            heading_row_color=ft.colors.with_opacity(1, "#f36229"),
            heading_text_style=ft.TextStyle(color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
            border=ft.BorderSide(color=ft.colors.BLACK, width=1),
            width=800,
            column_spacing=0,
            show_bottom_border=False,
            show_checkbox_column=False,
        )

        ### controles de paginação ###
        paginacao_controls = ft.Row(
            controls=[
                ft.IconButton(
                    icon=ft.icons.ARROW_BACK_IOS,
                    icon_color="black",
                    data="prev",
                    on_click=handle_page_change,
                    disabled=pagina_atual == 1
                ),
                ft.Text(f"{pagina_atual} de {total_paginas}"),
                ft.IconButton(
                    icon=ft.icons.ARROW_FORWARD_IOS,
                    icon_color="black",
                    data="next",
                    on_click=handle_page_change,
                    disabled=pagina_atual == total_paginas
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=20
        )

        ### atualiza o container com tabela e paginação ###
        cadastros_container.content = ft.Column(
            [tabela, paginacao_controls],
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=20
        )
        page.update()

    ### função para limpar filtros ###
    def limpar_filtros(e):
        nonlocal mes_selecionado, ano_selecionado, pagina_atual
        mes_selecionado = None
        ano_selecionado = None
        pagina_atual = 1
        mes_dropdown.value = None
        ano_dropdown.value = None
        mostrar_cadastros()

    ### função para deletar um estorno da planilha ###
    def deletar_estorno(row):
        try:
            caminho_arquivo = "banco/ProvisaoBD.xlsx"
            wb = load_workbook(caminho_arquivo)
            sheet = wb["Estornos"]
            
            chave_para_deletar = str(row['CHAVE']).strip()
            valor_estornado_para_deletar = row['VALOR ESTORNADO']

            if isinstance(valor_estornado_para_deletar, str):
                valor_estornado_para_deletar = valor_estornado_para_deletar.replace('R$', '').replace('.', '').replace(',', '.').strip()
                valor_estornado_para_deletar = float(valor_estornado_para_deletar)
            
            linha_para_deletar = None

            for i, excel_row in enumerate(sheet.iter_rows(values_only=True), start=1):
                chave_excel = str(excel_row[0]).strip() if excel_row[0] is not None else None
                valor_estornado_excel = excel_row[2]
                
                if isinstance(valor_estornado_excel, str):
                    if valor_estornado_excel.upper() == "VALOR ESTORNADO":
                        continue
                    valor_estornado_excel = valor_estornado_excel.replace('R$', '').replace('.', '').replace(',', '.').strip()
                    valor_estornado_excel = float(valor_estornado_excel)

                if chave_excel == chave_para_deletar and valor_estornado_excel == valor_estornado_para_deletar:
                    linha_para_deletar = i
                    break

            if linha_para_deletar:
                sheet.delete_rows(linha_para_deletar)
                wb.save(caminho_arquivo)

                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Estorno com chave {chave_para_deletar} e valor estornado R$ {valor_estornado_para_deletar:.2f} deletado com sucesso!")))
                mostrar_cadastros()

            else:
                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro: Estorno com chave {chave_para_deletar} e valor estornado R$ {valor_estornado_para_deletar:.2f} não encontrado!")))

        except Exception as e:
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro ao deletar o estorno: {str(e)}")))

    ### função para visualizar detalhes de um estorno ###
    def visualizar_estorno(row):
        caminho_arquivo = "banco/ProvisaoBD.xlsx"
        wb = load_workbook(caminho_arquivo)
        provisoes_sheet = wb["Provisões"]

        chave_provisao = str(row['CHAVE']).strip()

        dados_completos = None
        for linha in provisoes_sheet.iter_rows(values_only=True):
            chave_excel = str(linha[5]).strip() if linha[5] is not None else None
            if chave_excel == chave_provisao:
                dados_completos = {
                    'CHAVE': linha[5],
                    'CLIENTE': linha[1],
                    'RECEITA_BRUTA': linha[8],
                    'OBSERVACAO': linha[15]
                }
                break

        if dados_completos is None:
            page.show_snack_bar(ft.SnackBar(content=ft.Text("Dados não encontrados para esta CHAVE!")))
            return

        tabela = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Container(content=ft.Text("CHAVE", text_align=ft.TextAlign.CENTER, width=125))),
                ft.DataColumn(ft.Container(content=ft.Text("CLIENTE", text_align=ft.TextAlign.CENTER, width=300))),
                ft.DataColumn(ft.Container(content=ft.Text("RECEITA BRUTA", text_align=ft.TextAlign.CENTER, width=125))),
            ],
            rows=[
                ft.DataRow(cells=[
                    ft.DataCell(ft.Container(content=ft.Text(str(dados_completos['CHAVE']), text_align=ft.TextAlign.CENTER, width=125))),
                    ft.DataCell(ft.Container(content=ft.Text(dados_completos['CLIENTE'], text_align=ft.TextAlign.CENTER, width=300))),
                    ft.DataCell(ft.Container(content=ft.Text(f"R$ {float(dados_completos['RECEITA_BRUTA']):,.2f}".replace(",", "v").replace(".", ",").replace("v", "."), text_align=ft.TextAlign.CENTER, width=125))),
                ]),
            ],
            heading_row_color=ft.colors.with_opacity(1, "#212A57"),
            heading_text_style=ft.TextStyle(color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
            border=ft.BorderSide(color=ft.colors.BLACK, width=1),
            column_spacing=0,
            show_bottom_border=False,
            show_checkbox_column=False,
        )

        form_layout = ft.Column(
            controls=[
                tabela,
                ft.Text("OBSERVAÇÃO:", weight="bold"),
                ft.Text(dados_completos['OBSERVACAO'] if dados_completos['OBSERVACAO'] else "Sem observações."),
            ],
            alignment=ft.MainAxisAlignment.START,
            width=600,
            height=200,
        )

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Detalhes da Provisão (Chave {chave_provisao})", weight="bold"),
            content=form_layout,
            actions=[
                ft.Column(col={"sm": 2}, controls=[ft.ElevatedButton(
                    text="Fechar",
                    on_click=lambda e: fechar_modal(dialog),
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.with_opacity(1, "#212A57"),
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    width=145,
                    height=50,
                )]),
            ],
        )

        page.dialog = dialog
        dialog.open = True
        page.update()

    ### função para fechar o modal ###
    def fechar_modal(dialog):
        dialog.open = False
        page.update()

    ### função para abrir o formulário de cadastro de estorno ###
    def abrir_cadastro(e):
        dialog_content = criar_formulario_estorno(page)

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cadastro de Estornos", weight="bold"),
            content=dialog_content
        )

        page.dialog = dialog
        dialog.open = True
        page.update()

    ### função para abrir o FilePicker para importar planilha ###
    def importar_planilha(e):
        def file_picker_callback(result):
            if result is not None and result.files:
                arquivo_selecionado = result.files[0].path
                processar_importacao(arquivo_selecionado)

        file_picker.on_result = file_picker_callback
        file_picker.pick_files(allow_multiple=False, allowed_extensions=["xlsx"])

    ### função para processar o arquivo importado e salvar os dados no Excel ###
    def processar_importacao(arquivo):
        wb = load_workbook("banco/ProvisaoBD.xlsx")
        estornos_sheet = wb["Estornos"]

        estorno_df = pd.read_excel(arquivo, sheet_name='Estorno', header=0, dtype={"CHAVE": str})
        estorno_df.columns = ['CHAVE', 'DATA ESTORNO', 'VALOR ESTORNADO']
        estorno_df['DATA ESTORNO'] = pd.to_datetime(estorno_df['DATA ESTORNO'], errors='coerce')

        for r in dataframe_to_rows(estorno_df, index=False, header=False):
            estornos_sheet.append(r)

        formato_data = 'DD/MM/YYYY'
        for row in estornos_sheet.iter_rows(min_row=estornos_sheet.max_row - len(estorno_df) + 1, max_row=estornos_sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = formato_data

        for row in estornos_sheet.iter_rows(min_row=estornos_sheet.max_row - len(estorno_df) + 1, max_row=estornos_sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '@'

        col_idx_estorno_valor = estorno_df.columns.get_loc('VALOR ESTORNADO') + 1
        for row in estornos_sheet.iter_rows(min_row=estornos_sheet.max_row - len(estorno_df) + 1, max_row=estornos_sheet.max_row, min_col=col_idx_estorno_valor, max_col=col_idx_estorno_valor):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = r'_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

        wb.save("banco/ProvisaoBD.xlsx")

        page.show_snack_bar(ft.SnackBar(content=ft.Text("Importação concluída com sucesso!")))
        mostrar_cadastros()

    ### dropdown de mês ###
    mes_dropdown = ft.Dropdown(
        hint_text="Mês",
        options=[ft.dropdown.Option(mes) for mes in meses],
        on_change=handle_mes_change,
        width=140,
    )

    ### dropdown de ano ###
    ano_dropdown = ft.Dropdown(
        hint_text="Ano",
        options=[ft.dropdown.Option(ano) for ano in anos],
        on_change=handle_ano_change,
        width=140,
    )

    ### linha de botões superiores: limpar, cadastrar, importar ###
    botoes_superiores = ft.Row(
        controls=[
            ft.ElevatedButton(
                text="Limpar Filtros",
                on_click=limpar_filtros,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            ),
            ft.ElevatedButton(
                text="Cadastrar",
                on_click=abrir_cadastro,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            ),
            ft.ElevatedButton(
                text="Importar",
                on_click=importar_planilha,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            ),
        ],
        alignment=ft.MainAxisAlignment.END,
        spacing=10,
    )

    ### linha de filtros e botões ###
    filtros_row = ft.Row(
        controls=[
            mes_dropdown,
            ano_dropdown,
            botoes_superiores,
        ],
        alignment=ft.MainAxisAlignment.CENTER,
        spacing=20,
    )

    ### define o conteúdo do container ###
    container.content = ft.Column(
        controls=[
            filtros_row,
            cadastros_container,
        ],
        alignment=ft.MainAxisAlignment.START,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        spacing=20,
        width="100%"
    )

    mostrar_cadastros()
    page.update()
