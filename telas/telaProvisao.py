import flet as ft
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Alignment
import locale
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
from banco.banco import *
from funcoes.funcoes import fechar_modal
from formularios.cadProvisao import criar_formulario_provisao

### função principal da tela de provisão ###
def TelaProvisao(page, container):
    cadastros_container = ft.Container(padding=20)  ### contêiner para mostrar dados das provisões ###
    itens_por_pagina = 10  ### qtde de itens exibidos por página ###
    pagina_atual = 1  ### página atual iniciando como 1 ###
    mes_selecionado = None  ### mês inicial como nenhum selecionado ###
    ano_selecionado = None  ### ano inicial como nenhum selecionado ###
    cliente_selecionado = None  ### cliente inicial como nenhum selecionado ###
    total_paginas = 1  ### total de páginas inicial como 1 ###

    ### lista de meses para dropdown ###
    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]

    ### lista de anos para dropdown ###
    anos = [str(ano) for ano in range(2023, 2027)]

    ### inicializando o FilePicker para importar planilhas ###
    file_picker = ft.FilePicker(on_result=None)
    page.overlay.append(file_picker)

    ### função para determinar a cor do ícone baseado no valor estornado ###
    def determinar_cor_estorno(row):
        ### pegar chave da provisão e garantir que seja string ###
        chave_atual = str(int(row['CHAVE'])) if isinstance(row['CHAVE'], float) else str(row['CHAVE'])
        
        ### carregar dados de estorno da planilha ###
        estornos_df = pd.read_excel("banco/ProvisaoBD.xlsx", sheet_name='Estornos')

        ### converter chave para string para garantir compatibilidade ###
        estornos_df['CHAVE'] = estornos_df['CHAVE'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))

        ### filtrar estornos com base na chave ###
        estornos_filtrados = estornos_df[estornos_df['CHAVE'] == chave_atual]

        ### somar o total estornado ###
        total_estornado = estornos_filtrados['VALOR ESTORNADO'].sum()

        ### converter receita bruta para float ###
        receita_bruta = row['RECEITA BRUTA']
        if isinstance(receita_bruta, str):
            receita_bruta = receita_bruta.replace('.', '').replace(',', '.')
            receita_bruta = float(receita_bruta)

        ### calcular o valor que falta estornar ###
        faltar_estornar = receita_bruta - total_estornado

        ### definir cor com base no quanto foi estornado ###
        if total_estornado == 0:
            return ft.colors.RED  ### nada estornado ###
        elif faltar_estornar > 0:
            return ft.colors.AMBER  ### estorno parcial ###
        else:
            return ft.colors.GREEN  ### estorno total ###

    ### funções para manipular a mudança de mês, ano, e cliente ###
    def handle_mes_change(e):
        nonlocal mes_selecionado, pagina_atual
        mes_selecionado = meses.index(e.control.value) + 1 if e.control.value else None
        pagina_atual = 1
        mostrar_cadastros()
    
    def format_currency(value):
        ### função para formatar valores como moeda ###
        return f"R$ {value:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")

    def handle_ano_change(e):
        nonlocal ano_selecionado, pagina_atual
        ano_selecionado = int(e.control.value) if e.control.value else None
        pagina_atual = 1
        mostrar_cadastros()

    def dropdown_changed(e):
        nonlocal cliente_selecionado, pagina_atual
        cliente_selecionado = e.control.value if e.control.value else None
        pagina_atual = 1
        mostrar_cadastros()

    ### função para manipular mudança de página na tabela ###
    def handle_page_change(e):
        nonlocal pagina_atual
        if e.control.data == "prev" and pagina_atual > 1:
            pagina_atual -= 1
        elif e.control.data == "next" and pagina_atual < total_paginas:
            pagina_atual += 1
        mostrar_cadastros()

    ### função que exibe os cadastros filtrados com paginação ###
    def mostrar_cadastros():
        nonlocal total_paginas
        ### filtrar dados baseado em mês, ano e cliente ###
        dados = filtrar_dados_provisao(mes_selecionado, ano_selecionado, cliente_selecionado)
        total_registros = len(dados)
        total_paginas = (total_registros + itens_por_pagina - 1) // itens_por_pagina

        ### paginar os dados ###
        inicio = (pagina_atual - 1) * itens_por_pagina
        fim = inicio + itens_por_pagina
        dados_pagina = dados.iloc[inicio:fim]

        tabela_dados = []
        ### loop para adicionar dados à tabela ###
        for _, row in dados_pagina.iterrows():
            tabela_dados.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Container(
                            content=ft.Text(str(row['CHAVE']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=165
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(row['CLIENTE'], text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=205
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(row['TIPO DOC'], text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=140
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(row['DATA PROVISÃO'].strftime('%d/%m/%Y'), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=150
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(format_currency(row['RECEITA BRUTA']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=210
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.CIRCLE,
                                icon_color=determinar_cor_estorno(row),
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.DELETE,
                                icon_color="#212a57",
                                on_click=lambda e, row=row: deletar_provisao(row)
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.IconButton(
                                icon=ft.icons.VISIBILITY,
                                icon_color="#212a57",
                                on_click=lambda e, row=row: visualizar_provisao(row)
                            ),
                            alignment=ft.alignment.center,
                            width=50
                        )),
                    ]
                )
            )

        ### criar tabela de dados com colunas e exibir ###
        tabela = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Container(
                    content=ft.Text("CHAVE", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=165
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("CLIENTE", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=205
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("TIPO DOC", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=140
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("DATA PROVISÃO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=150
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("RECEITA BRUTA", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=210
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("AÇÕES", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=50
                )),
            ],
            rows=tabela_dados,
            heading_row_color=ft.colors.with_opacity(1, "#f36229"),
            heading_text_style=ft.TextStyle(color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
            border=ft.BorderSide(color=ft.colors.BLACK, width=1),
            width=1080,
            column_spacing=0,
            show_bottom_border=False,
            show_checkbox_column=False,
        )

        ### controles de paginação ###
        paginacao_controls = ft.Row(
            controls=[
                ft.IconButton(
                    icon=ft.icons.ARROW_BACK_IOS,
                    icon_color="black",
                    data="prev",
                    on_click=handle_page_change,
                    disabled=pagina_atual == 1
                ),
                ft.Text(f"{pagina_atual} de {total_paginas}"),
                ft.IconButton(
                    icon=ft.icons.ARROW_FORWARD_IOS,
                    icon_color="black",
                    data="next",
                    on_click=handle_page_change,
                    disabled=pagina_atual == total_paginas
                ),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=20
        )

        ### atualizar contêiner com tabela e paginação ###
        cadastros_container.content = ft.Column(
            [tabela, paginacao_controls],
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=20
        )
        page.update()

    ### função para limpar filtros de pesquisa ###
    def limpar_filtros(e):
        nonlocal mes_selecionado, ano_selecionado, cliente_selecionado, pagina_atual
        mes_selecionado = None
        ano_selecionado = None
        cliente_selecionado = None
        pagina_atual = 1
        mes_dropdown.value = None
        ano_dropdown.value = None
        cliente_dropdown.value = None
        mostrar_cadastros()

    ### função para deletar uma provisão da planilha ###
    def deletar_provisao(row):
        try:
            ### carregar arquivo Excel ###
            caminho_arquivo = "banco/ProvisaoBD.xlsx"
            wb = load_workbook(caminho_arquivo)
            sheet = wb["Provisões"]
            
            ### procurar linha com a chave correspondente ###
            chave_para_deletar = str(row['CHAVE']).strip()
            linha_para_deletar = None

            for i, excel_row in enumerate(sheet.iter_rows(values_only=True), start=1):
                chave_excel = str(excel_row[5]).strip() if excel_row[5] is not None else None
                if chave_excel == chave_para_deletar:
                    linha_para_deletar = i
                    break
            
            ### deletar a linha encontrada ###
            if linha_para_deletar:
                sheet.delete_rows(linha_para_deletar)
                wb.save(caminho_arquivo)

                ### exibir mensagem de sucesso ###
                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Provisão {chave_para_deletar} deletada com sucesso!")))

                ### atualizar tabela ###
                mostrar_cadastros()

            else:
                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro: Provisão com chave {chave_para_deletar} não encontrada!")))

        except Exception as e:
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro ao deletar a provisão: {str(e)}")))

    ### função para visualizar uma provisão em um modal ###
    def visualizar_provisao(row):

        dados_completos = obter_dados_provisao(row['CHAVE'])
        if dados_completos is None:
            page.show_snack_bar(ft.SnackBar(content=ft.Text("Dados não encontrados para esta CHAVE!")))
            return
        
        ### Função para copiar o valor da CHAVE para a área de transferência ###
        def copiar_para_area_transferencia(e):
            page.set_clipboard(chave_atual)
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Chave {chave_atual} copiada!")))

        ### Garantir que 'DATA PROVISÃO' seja um objeto datetime antes de aplicar 'strftime' ###
        data_provisao = dados_completos['DATA PROVISÃO']
        
        if isinstance(data_provisao, str):
            try:
                data_provisao = pd.to_datetime(data_provisao, errors='coerce')
            except ValueError:
                data_provisao = None

        ### Função para garantir formatação correta de valores monetários ###
        def formatar_monetario(valor):
            try:
                if isinstance(valor, (float, int)):
                    return f"R$ {valor:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
                if isinstance(valor, str):
                    valor = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
                    return f"R$ {float(valor):,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
            except (ValueError, TypeError):
                return "Valor inválido"

        ### Obter valores monetários formatados ###
        receita_bruta = formatar_monetario(dados_completos['RECEITA BRUTA'])
        receita_liquida = formatar_monetario(dados_completos['RECEITA LÍQUIDA'])
        icms = formatar_monetario(dados_completos['ICMS'])
        iss = formatar_monetario(dados_completos['ISS'])
        pis = formatar_monetario(dados_completos['PIS'])
        cofins = formatar_monetario(dados_completos['COFINS'])
        cprb = formatar_monetario(dados_completos['CPRB'])
        
        ### Converter RECEITA BRUTA para float caso ainda seja string ###
        receita_bruta_valor = dados_completos['RECEITA BRUTA']
        if isinstance(receita_bruta_valor, str):
            receita_bruta_valor = receita_bruta_valor.replace('.', '').replace(',', '.')
            try:
                receita_bruta_valor = float(receita_bruta_valor)
            except ValueError:
                receita_bruta_valor = 0.0  # Caso a conversão falhe, usar 0 como fallback

        ### Carregar estornos baseados na CHAVE ###
        estornos_df = pd.read_excel("banco/ProvisaoBD.xlsx", sheet_name='Estornos')

        ### Converter a coluna de datas de estornos diretamente para datetime ###
        estornos_df['DATA ESTORNO'] = pd.to_datetime(estornos_df['DATA ESTORNO'], errors='coerce')

        ### Converter chave para string ###
        estornos_df['CHAVE'] = estornos_df['CHAVE'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))
        chave_atual = str(int(dados_completos['CHAVE'])) if isinstance(dados_completos['CHAVE'], float) else str(dados_completos['CHAVE'])

        ### Filtrar estornos com base na CHAVE ###
        estornos_filtrados = estornos_df[estornos_df['CHAVE'] == chave_atual]

        ### Calcular total estornado ###
        total_estornado = estornos_filtrados['VALOR ESTORNADO'].sum()

        ### Calcular o valor que falta estornar ###
        faltar_estornar = receita_bruta_valor - total_estornado
        faltar_estornar_formatado = formatar_monetario(faltar_estornar)

        ### Criar DataTable para exibir os estornos ###
        tabela_dados = []
        for _, estorno_row in estornos_filtrados.iterrows():
            data_estorno = estorno_row['DATA ESTORNO']
            if pd.notna(data_estorno):  # Verifica se a data não é nula
                data_estorno_formatada = data_estorno.strftime('%d/%m/%Y')
            else:
                data_estorno_formatada = "Data inválida"

            tabela_dados.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Container(
                            content=ft.Text(str(estorno_row['CHAVE']), text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=175
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(data_estorno_formatada, text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=175
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(f"R$ {estorno_row['VALOR ESTORNADO']:,.2f}", text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            width=300
                        )),
                    ]
                )
            )

        tabela = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Container(
                    content=ft.Text("CHAVE", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=175
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("DATA ESTORNO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=175
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("VALOR ESTORNADO", text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    width=300
                )),
            ],
            rows=tabela_dados,
            heading_row_color=ft.colors.with_opacity(1, "#212A57"),
            heading_text_style=ft.TextStyle(color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
            border=ft.BorderSide(color=ft.colors.BLACK, width=1),
            width=650,
            height=250,
            column_spacing=0,
            show_bottom_border=False,
            show_checkbox_column=False,
        )

        ### Layout do modal com as informações completas da provisão ###
        form_layout = ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Text("DATA PROVISÃO:", weight="bold"),
                        ft.Text(data_provisao.strftime('%d/%m/%Y') if data_provisao else "Data inválida"),
                        ft.Text("CLIENTE:", weight="bold"),
                        ft.Text(dados_completos['CLIENTE']),
                    ]
                ),
                ft.Row(
                    controls=[
                        ft.Text("CLASSIFICAÇÃO:", weight="bold"),
                        ft.Text(dados_completos['CLASSIF']),
                        ft.Text("CHAVE:", weight="bold"),
                        ft.Text(dados_completos['CHAVE']),
                        ft.IconButton(
                        icon=ft.icons.COPY,
                        tooltip="Copiar",
                        on_click=copiar_para_area_transferencia,
                        icon_size=20
                        ),
                        ft.Text("NUM DOC:", weight="bold"),
                        ft.Text(dados_completos['NUM DOC']),
                        ft.Text("TIPO DOC:", weight="bold"),
                        ft.Text(dados_completos['TIPO DOC'])
                    ]
                ),
                ft.Row(
                    controls=[
                        ft.Text("UND NEGÓCIO:", weight="bold"),
                        ft.Text(dados_completos['UND NEGÓCIO']),
                        ft.Text("RECEITA BRUTA:", weight="bold"),
                        ft.Text(receita_bruta),
                        ft.Text("RECEITA LÍQUIDA:", weight="bold"),
                        ft.Text(receita_liquida)
                    ]
                ),
                ft.Row(
                    controls=[
                        ft.Text("ICMS:", weight="bold"),
                        ft.Text(icms),
                        ft.Text("ISS:", weight="bold"),
                        ft.Text(iss),
                        ft.Text("PIS:", weight="bold"),
                        ft.Text(pis),
                        ft.Text("COFINS:", weight="bold"),
                        ft.Text(cofins),
                        ft.Text("CPRB:", weight="bold"),
                        ft.Text(cprb)
                    ]
                ),
                ft.Row(
                    controls=[
                        ft.Text("OBSERVAÇÃO:", weight="bold"),
                        ft.Text(dados_completos['OBSERVACAO'])
                    ]
                ),
                ft.Row(
                    controls=[
                        ft.Text(f"FALTA ESTORNAR: {faltar_estornar_formatado}", weight="bold"),
                    ]
                ),
                tabela,  ### Adicionar DataTable com estornos ###
            ],
            alignment=ft.MainAxisAlignment.START,
        )

        ### Exibir modal com os dados ###
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Provisão e Estorno da Chave {chave_atual}", weight="bold"),
            content=form_layout,
            actions=[
                ft.Column(col={"sm": 2}, controls=[ft.ElevatedButton(
                    text="Fechar",
                    on_click=lambda e: fechar_modal(dialog),
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.with_opacity(1, "#212A57"),
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    width=145,
                    height=50,
                )]),
            ],
        )

        page.dialog = dialog
        dialog.open = True
        page.update()


    ### função para fechar o modal ###
    def fechar_modal(dialog):
        dialog.open = False
        page.update()

    ### função para buscar os dados completos de uma provisão ###
    def obter_dados_provisao(chave):
        wb = load_workbook("banco/ProvisaoBD.xlsx")
        provisoes_sheet = wb["Provisões"]

        chave = str(chave).strip()  ### converter chave para string ###
        for row in provisoes_sheet.iter_rows(values_only=True):
            chave_row = row[5]  ### coluna da chave ###
            if isinstance(chave_row, float):
                chave_row = str(int(chave_row))
            elif chave_row is not None:
                chave_row = str(chave_row).strip()

            if chave_row == chave:
                dados_completos = {
                    'DATA PROVISÃO': row[0],
                    'CLIENTE': row[1],
                    'UND NEGÓCIO': row[2],
                    'I.C': row[3],
                    'TIPO DOC': row[4],
                    'CHAVE': row[5],
                    'NUM DOC': row[6],
                    'CLASSIF': row[7],
                    'RECEITA BRUTA': row[8],
                    'ICMS': row[9],
                    'ISS': row[10],
                    'PIS': row[11],
                    'COFINS': row[12],
                    'CPRB': row[13],
                    'RECEITA LÍQUIDA': row[14],
                    'OBSERVACAO': row[15]
                }
                return dados_completos

        return None

    ### função para abrir o formulário de cadastro em um modal ###
    def abrir_cadastro(e):
        dialog_content = criar_formulario_provisao(page)

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cadastro de Provisões", weight="bold"),
            content=dialog_content
        )

        page.dialog = dialog
        dialog.open = True
        page.update()

    ### função para abrir o file picker e importar planilha ###
    def importar_planilha(e):
        def file_picker_callback(result):
            if result is not None and result.files:
                arquivo_selecionado = result.files[0].path
                processar_importacao(arquivo_selecionado)

        file_picker.on_result = file_picker_callback
        file_picker.pick_files(allow_multiple=False, allowed_extensions=["xlsx"])

    ### função para processar o arquivo importado ###
    def processar_importacao(arquivo):
        wb_banco = load_workbook("banco/ProvisaoBD.xlsx")
        provisoes_sheet = wb_banco["Provisões"]
        estornos_sheet = wb_banco["Estornos"]

        wb = load_workbook(arquivo, data_only=True)
        provisao_df = pd.read_excel(arquivo, sheet_name='Provisão', header=0, dtype={"CHAVE": str, "UND NEGÓCIO": str, "I.C": str})
        estorno_df = pd.read_excel(arquivo, sheet_name='Estorno', header=0, dtype={"CHAVE": str})

        if len(provisao_df.columns) == 17:
            provisao_df.columns = [
                'DATA PROVISÃO', 'CLIENTE', 'UND NEGÓCIO', 'I.C', 'TIPO DOC',
                'CHAVE', 'NUM DOC', 'CLASSIF', 'RECEITA BRUTA', 'ICMS', 'ISS',
                'PIS', 'COFINS', 'CPRB', 'RECEITA LÍQUIDA', 'OBSERVACAO', 'EXTRA'
            ]
            provisao_df.drop(columns=['EXTRA'], inplace=True)
        elif len(provisao_df.columns) == 16:
            provisao_df.columns = [
                'DATA PROVISÃO', 'CLIENTE', 'UND NEGÓCIO', 'I.C', 'TIPO DOC',
                'CHAVE', 'NUM DOC', 'CLASSIF', 'RECEITA BRUTA', 'ICMS', 'ISS',
                'PIS', 'COFINS', 'CPRB', 'RECEITA LÍQUIDA', 'OBSERVACAO'
            ]

        estorno_df.columns = [
            'CHAVE', 'DATA ESTORNO', 'VALOR ESTORNADO'
        ]

        ### converter colunas de data para datetime ###
        provisao_df['DATA PROVISÃO'] = pd.to_datetime(provisao_df['DATA PROVISÃO'], errors='coerce')
        estorno_df['DATA ESTORNO'] = pd.to_datetime(estorno_df['DATA ESTORNO'], errors='coerce')

        ### adicionar novos dados à planilha ###
        for r in dataframe_to_rows(provisao_df, index=False, header=False):
            provisoes_sheet.append(r)
        
        for r in dataframe_to_rows(estorno_df, index=False, header=False):
            estornos_sheet.append(r)

        ### formatar colunas de data ###
        formato_data = 'DD/MM/YYYY'
        for row in provisoes_sheet.iter_rows(min_row=provisoes_sheet.max_row - len(provisao_df) + 1, max_row=provisoes_sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = formato_data

        for row in estornos_sheet.iter_rows(min_row=estornos_sheet.max_row - len(estorno_df) + 1, max_row=estornos_sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = formato_data

        ### formatar colunas como texto ###
        for row in provisoes_sheet.iter_rows(min_row=provisoes_sheet.max_row - len(provisao_df) + 1, max_row=provisoes_sheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = '@'

        for row in provisoes_sheet.iter_rows(min_row=provisoes_sheet.max_row - len(provisao_df) + 1, max_row=provisoes_sheet.max_row, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = '@'

        for row in estornos_sheet.iter_rows(min_row=estornos_sheet.max_row - len(estorno_df) + 1, max_row=estornos_sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '@'

        ### aplicar formatação monetária ###
        cols_monetarios_provisoes = ['RECEITA BRUTA', 'RECEITA LÍQUIDA', 'ICMS', 'ISS', 'PIS', 'COFINS', 'CPRB']
        formato_moeda = r'_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
        
        for col_name in cols_monetarios_provisoes:
            col_idx = provisao_df.columns.get_loc(col_name) + 1
            for row in provisoes_sheet.iter_rows(min_row=provisoes_sheet.max_row - len(provisao_df) + 1, max_row=provisoes_sheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = formato_moeda

        col_idx_estorno_valor = estorno_df.columns.get_loc('VALOR ESTORNADO') + 1
        for row in estornos_sheet.iter_rows(min_row=estornos_sheet.max_row - len(estorno_df) + 1, max_row=estornos_sheet.max_row, min_col=col_idx_estorno_valor, max_col=col_idx_estorno_valor):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = formato_moeda

        wb_banco.save("banco/ProvisaoBD.xlsx")

        page.show_snack_bar(ft.SnackBar(content=ft.Text("Importação concluída com sucesso!")))
        mostrar_cadastros()

    ### dropdowns para filtros de mês, ano e cliente ###
    cliente_dropdown = ft.Dropdown(
        hint_text="Cliente",
        options=[ft.dropdown.Option(cliente) for cliente in sorted(obter_clientes())],
        on_change=dropdown_changed,
        width=290,
    )

    mes_dropdown = ft.Dropdown(
        hint_text="Mês",
        options=[ft.dropdown.Option(mes) for mes in meses],
        on_change=handle_mes_change,
        width=140,
    )

    ano_dropdown = ft.Dropdown(
        hint_text="Ano",
        options=[ft.dropdown.Option(ano) for ano in anos],
        on_change=handle_ano_change,
        width=140,
    )

    ### botões para limpar filtros, cadastrar e importar planilhas ###
    botoes_superiores = ft.Row(
        controls=[
            ft.ElevatedButton(
                text="Limpar Filtros",
                on_click=limpar_filtros,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            ),
            ft.ElevatedButton(
                text="Cadastrar",
                on_click=abrir_cadastro,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            ),
            ft.ElevatedButton(
                text="Importar",
                on_click=importar_planilha,
                style=ft.ButtonStyle(
                    bgcolor=ft.colors.with_opacity(1, "#f36229"),
                    color=ft.colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=8),
                ),
                width=145,
                height=50,
            ),
        ],
        alignment=ft.MainAxisAlignment.END,
        spacing=10,
    )

    ### linha com filtros e botões superiores ###
    filtros_row = ft.Row(
        controls=[
            mes_dropdown,
            ano_dropdown,
            cliente_dropdown,
            botoes_superiores,
        ],
        alignment=ft.MainAxisAlignment.CENTER,
        spacing=20,
    )

    ### adicionar tudo ao container e exibir ###
    container.content = ft.Column(
        controls=[
            filtros_row,
            cadastros_container,
        ],
        alignment=ft.MainAxisAlignment.START,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        spacing=20,
        width="100%"
    )

    mostrar_cadastros()
    page.update()
