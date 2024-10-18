import flet as ft
from banco.banco import obter_clientes, carregar_impostos_de_json, salvar_dados_excel
from funcoes.funcoes import format_currency, arquivo_selecionado
from openpyxl import load_workbook
import locale

# Configurando a localização para português do Brasil
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

def criar_formulario_estorno(page):
    ### Função para salvar os dados no Excel ###
    def salvar_dados(e):
        try:
            # Carregar a planilha existente e a sheet "Estornos"
            caminho_planilha = "banco/ProvisaoBD.xlsx"
            planilha = load_workbook(caminho_planilha)
            sheet_provisao = planilha["Provisões"]
            sheet_estornos = planilha["Estornos"]

            # Obter a chave e o valor estornado
            chave = chave_field.value
            valor_estornado = float(valor_estornado_field.value.replace("R$", "").strip())

            # Filtrar a receita bruta da chave informada na planilha Provisões
            receita_bruta = 0
            for row in sheet_provisao.iter_rows(min_row=2, values_only=True):
                if str(row[5]) == chave:  # A coluna 5 é a 'CHAVE'
                    receita_bruta = float(row[8])  # Coluna 7 é a 'RECEITA BRUTA'
                    break

            # Somar os valores estornados já registrados para a mesma chave
            total_estornado = 0
            for row in sheet_estornos.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == chave:  # Coluna 0 é a 'CHAVE'
                    total_estornado += float(row[2])  # Coluna 2 é o 'VALOR ESTORNADO'

            # Calcular o valor restante possível de estorno
            valor_restante = receita_bruta - total_estornado

            # Verificar se o valor estornado informado é maior que o permitido
            if valor_estornado > valor_restante:
                # Limpar os campos, fechar o modal e exibir mensagem de erro
                chave_field.value = ""
                meses_dropdown.value = None
                anos_dropdown.value = None
                valor_estornado_field.value = ""
                page.update()

                fechar_modal(page)  # Fechar o modal
                page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Valor escolhido é maior que o possível para efetuar estorno ({format_currency(valor_restante)})."), bgcolor=ft.colors.RED))
                return

            # Criar a data com o dia 15 do mês e ano selecionado
            mes_selecionado = meses_dropdown.value
            ano_selecionado = anos_dropdown.value
            dia_fixo = 15
            data_estorno_formatada = f"15/{mes_selecionado}/{ano_selecionado}"

            # Adicionar os dados à próxima linha vazia
            nova_linha = [
                chave,
                data_estorno_formatada,
                valor_estornado
            ]

            # Inserir a nova linha na sheet "Estornos"
            sheet_estornos.append(nova_linha)

            # Salvar a planilha atualizada
            planilha.save(caminho_planilha)

            # Exibir mensagem de sucesso
            page.show_snack_bar(ft.SnackBar(content=ft.Text("Estorno salvo com sucesso!"), bgcolor=ft.colors.GREEN))

        except Exception as ex:
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro ao salvar o estorno: {ex}"), bgcolor=ft.colors.RED))

    ### Criação dos componentes do formulário ###
    chave_field = ft.TextField(
        label="Chave",
        hint_text="Digite a chave",
        width="100%"
    )

    # Dropdown para selecionar o mês
    meses_dropdown = ft.Dropdown(
        label="Mês Estorno",
        options=[
            ft.dropdown.Option("01", "Janeiro"),
            ft.dropdown.Option("02", "Fevereiro"),
            ft.dropdown.Option("03", "Março"),
            ft.dropdown.Option("04", "Abril"),
            ft.dropdown.Option("05", "Maio"),
            ft.dropdown.Option("06", "Junho"),
            ft.dropdown.Option("07", "Julho"),
            ft.dropdown.Option("08", "Agosto"),
            ft.dropdown.Option("09", "Setembro"),
            ft.dropdown.Option("10", "Outubro"),
            ft.dropdown.Option("11", "Novembro"),
            ft.dropdown.Option("12", "Dezembro")
        ],
        width="100%"
    )

    # Dropdown para selecionar o ano
    anos_dropdown = ft.Dropdown(
        label="Ano Estorno",
        options=[
            ft.dropdown.Option("2024"),
            ft.dropdown.Option("2025"),
            ft.dropdown.Option("2026")
        ],
        width="100%"
    )

    valor_estornado_field = ft.TextField(
        label="Valor Estornado",
        hint_text="Digite o valor estornado",
        width="100%",
    )

    ### Layout do formulário usando ResponsiveRow ###
    form_layout = ft.Column(
        controls=[
            ft.ResponsiveRow([
                ft.Column(col={"sm": 6}, controls=[chave_field]),
                ft.Column(col={"sm": 6}, controls=[valor_estornado_field]),
            ]),
            ft.ResponsiveRow([
                ft.Column(col={"sm": 6}, controls=[meses_dropdown]),
                ft.Column(col={"sm": 6}, controls=[anos_dropdown]),
            ]),
            ft.ResponsiveRow([
                ft.Column(col={"sm": 3}, controls=[ft.ElevatedButton(
                    text="Salvar",
                    on_click=salvar_dados,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.with_opacity(1, "#f36229"),
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    width=145,
                    height=50,
                ),]),
                ft.Column(col={"sm": 3}, controls=[ft.ElevatedButton(
                    text="Fechar",
                    on_click=lambda e: fechar_modal(page),
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.with_opacity(1, "#f36229"),
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    width=145,
                    height=50,
                )]),
            ],
            alignment=ft.MainAxisAlignment.START,
            spacing=10,
            ),
        ],
        alignment=ft.MainAxisAlignment.CENTER,
        spacing=20,
        width=400,
        height=200,
    )

    return form_layout

### Função para fechar o modal ###
def fechar_modal(page):
    if page.dialog:
        page.dialog.open = False
        page.dialog = None
        page.update()
