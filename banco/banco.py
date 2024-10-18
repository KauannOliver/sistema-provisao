import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, Font
from openpyxl.utils import get_column_letter
import locale
import json

### Função para carregar os dados da planilha Excel ###
def carregar_dados(caminho_arquivo, sheet_name='Provisões'):
    df = pd.read_excel(caminho_arquivo, sheet_name=sheet_name)
    return df

# Certifique-se de definir o local para português do Brasil
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

def salvar_dados_excel(caminho_planilha, nova_linha):
    try:
        # Carregar a planilha existente
        planilha = load_workbook(caminho_planilha)
        aba = planilha["Provisões"]

        # Verificar se os estilos já foram adicionados anteriormente
        if "estilo_texto" not in planilha.named_styles:
            estilo_texto = NamedStyle(name="estilo_texto", alignment=Alignment(horizontal="left"))
            planilha.add_named_style(estilo_texto)
        
        if "estilo_contabil" not in planilha.named_styles:
            estilo_contabil = NamedStyle(name="estilo_contabil", number_format='_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-')
            planilha.add_named_style(estilo_contabil)

        # Adicionar os dados à próxima linha vazia
        aba.append(nova_linha)
        linha_atual = aba.max_row

        # Verificar e ajustar a data na primeira coluna (DATA PROVISÃO)
        data_provisao = nova_linha[0]
        if isinstance(data_provisao, str):
            try:
                # Tentar converter o formato "mês por extenso/ano" para o formato "dia/mês/ano"
                data_provisao = datetime.strptime(data_provisao, '%B/%y').strftime('%d/%m/%Y')
            except ValueError:
                # Se falhar, tentar o formato padrão
                data_provisao = datetime.strptime(data_provisao, "%d/%m/%Y")
        aba.cell(row=linha_atual, column=1).value = data_provisao
        aba.cell(row=linha_atual, column=1).number_format = 'DD/MM/YYYY'

        # Definindo o índice fixo das colunas que contêm valores monetários e de texto
        cols_monetarios_provisoes = [9, 10, 11, 12, 13, 14, 15]  # Índices das colunas de valores monetários (R$)
        cols_texto = [2, 3, 4, 5, 6, 7, 8, 16]  # Índices das colunas de texto

        # Aplicar formatação de texto nas colunas de texto
        for col in cols_texto:
            aba.cell(row=linha_atual, column=col).number_format = '@'  # Definir formato de texto

        # Aplicar formatação de moeda nas colunas monetárias
        for col in cols_monetarios_provisoes:
            aba.cell(row=linha_atual, column=col).number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'

        # Salvar a planilha
        planilha.save(caminho_planilha)
        return "Cadastro salvo com sucesso!"
    except Exception as ex:
        return f"Erro ao salvar o cadastro: {ex}"

### Função para filtrar dados com base em mês, ano e cliente ###
def filtrar_dados_provisao(mes=None, ano=None, cliente=None):
    df = carregar_dados('banco/ProvisaoBD.xlsx', sheet_name='Provisões')
    
    ### Verifica e aplica os filtros de mês e ano no formato DD/MM/AAAA ###
    if mes is not None and ano is not None:
        df['DATA PROVISÃO'] = pd.to_datetime(df['DATA PROVISÃO'], format='%d/%m/%Y')
        df = df[(df['DATA PROVISÃO'].dt.month == mes) & (df['DATA PROVISÃO'].dt.year == ano)]

    if cliente:
        df = df[df['CLIENTE'] == cliente]

    return df

### Função para obter a lista de clientes única para o dropdown ###
def obter_clientes():
    df = carregar_dados('banco/ProvisaoBD.xlsx', sheet_name='Provisões')
    return df['CLIENTE'].unique().tolist()

### Função para carregar os impostos do arquivo JSON com codificação UTF-8 ###
def carregar_impostos_de_json():
    json_file = 'banco/impostos.json'  ### Caminho do arquivo JSON de impostos ###
    
    # Abrir o arquivo JSON com a codificação correta
    with open(json_file, 'r', encoding='utf-8') as f:
        impostos_dict = json.load(f)
    
    return impostos_dict
