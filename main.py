import flet as ft
import pandas as pd
import json
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from funcoes.funcoes import botao_menu_lateral
from telas.telaProvisao import TelaProvisao
from telas.telaEstorno import TelaEstorno
from telas.telaCliente import TelaCliente

def exportar_impostos_para_json():
    caminho_planilha = "banco/ProvisaoBD.xlsx"
    sheet_name = "Impostos"

    try:
        ### carrega a planilha ###
        df = pd.read_excel(caminho_planilha, sheet_name=sheet_name)

        ### processa os dados para o formato JSON ###
        impostos_dict = {}
        for _, row in df.iterrows():
            chave = str(int(row["I.C"])).strip()  ### converte "I.C" para string sem formatação adicional ###
            impostos_dict[chave] = {
                "CLIENTE": row["CLIENTE"],
                "UND NEGOCIO": int(row["UND NEGÓCIO"]),
                "ICMS": float(row["ICMS"]),
                "ISS": float(row["ISS"]),
                "PIS": float(row["PIS"]),
                "COFINS": float(row["COFINS"]),
                "CPRB": float(row["CPRB"])
            }

        ### caminho do arquivo JSON ###
        caminho_json = "banco/impostos.json"

        ### se o arquivo já existir, removê-lo ###
        if os.path.exists(caminho_json):
            os.remove(caminho_json)

        ### criar e salvar o novo arquivo JSON ###
        with open(caminho_json, "w", encoding="utf-8") as json_file:
            json.dump(impostos_dict, json_file, ensure_ascii=False, indent=4)

        print("Impostos exportados para JSON com sucesso!")

    except Exception as e:
        print(f"Erro ao exportar impostos para JSON: {e}")

def inserir_dados_no_excel():
    caminho_json = "banco/impostos.json"
    caminho_excel = "Modelo Importação.xlsx"
    sheet_name = "Impostos"

    try:
        ### verifica se o JSON existe ###
        if not os.path.exists(caminho_json):
            print("Arquivo impostos.json não encontrado.")
            return

        ### carrega os dados do JSON ###
        with open(caminho_json, "r", encoding="utf-8") as json_file:
            impostos_data = json.load(json_file)

        ### abre a planilha Modelo Importação V2.xlsx ###
        if os.path.exists(caminho_excel):
            wb = load_workbook(caminho_excel)
        else:
            ### se o arquivo não existir, cria um novo Workbook ###
            wb = Workbook()

        ### verifica se a sheet Impostos existe, caso contrário, cria uma nova ###
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)

        ### exclui todas as linhas a partir da linha 2 (mantém o cabeçalho) ###
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(2, max_row - 1)

        ### escreve os cabeçalhos das colunas ###
        sheet["A1"] = "CLIENTE"
        sheet["B1"] = "I.C"
        sheet["C1"] = "UND NEGOCIO"
        sheet["D1"] = "ICMS"
        sheet["E1"] = "ISS"
        sheet["F1"] = "PIS"
        sheet["G1"] = "COFINS"
        sheet["H1"] = "CPRB"

        ### preenche os dados do JSON na planilha Excel ###
        for ic, data in impostos_data.items():
            sheet.append([
                data["CLIENTE"],
                ic,  ### preencher o I.C diretamente do JSON ###
                data["UND NEGOCIO"],
                data["ICMS"],
                data["ISS"],
                data["PIS"],
                data["COFINS"],
                data["CPRB"]
            ])

        ### salva as mudanças no Excel ###
        wb.save(caminho_excel)
        print("Dados inseridos no Excel com sucesso!")

    except Exception as e:
        print(f"Erro ao inserir dados no Excel: {e}")

### função para baixar pendências ###
def baixar_pendencias():
    caminho_provisao = "banco/ProvisaoBD.xlsx"
    provisao_sheet = "Provisões"
    estornos_sheet = "Estornos"

    try:
        ### carregar as planilhas ###
        provisao_df = pd.read_excel(caminho_provisao, sheet_name=provisao_sheet)
        estornos_df = pd.read_excel(caminho_provisao, sheet_name=estornos_sheet)

        ### garantir que CHAVE seja string ###
        provisao_df['CHAVE'] = provisao_df['CHAVE'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))
        estornos_df['CHAVE'] = estornos_df['CHAVE'].apply(lambda x: str(int(x)) if isinstance(x, float) else str(x))

        ### criar uma nova coluna 'ESTORNOS' somando os valores da tabela de estornos ###
        estornos_somados = estornos_df.groupby('CHAVE')['VALOR ESTORNADO'].sum().reset_index()
        estornos_somados.columns = ['CHAVE', 'ESTORNOS']

        ### mesclar os estornos na tabela de provisões ###
        dados_completos = pd.merge(provisao_df, estornos_somados, on='CHAVE', how='left')
        dados_completos['ESTORNOS'] = dados_completos['ESTORNOS'].fillna(0)

        ### calcular pendências (Receita Bruta - Estornos) ###
        dados_completos['PENDENCIAS'] = dados_completos['RECEITA BRUTA'] - dados_completos['ESTORNOS']

        ### filtrar apenas as provisões com pendências diferentes de 0 ###
        dados_filtrados = dados_completos[dados_completos['PENDENCIAS'] != 0]

        ### verificar a presença de todas as colunas necessárias ###
        colunas_necessarias = [
            'DATA PROVISÃO', 'CLIENTE', 'UND NEGÓCIO', 'I.C', 'TIPO DOC', 'CHAVE', 'NÚM DOC', 'CLASSIFICAÇÃO',
            'RECEITA BRUTA', 'RECEITA LÍQUIDA', 'OBSERVAÇÃO'
        ]

        ### se houver colunas ausentes, preencher com valores padrão ###
        for coluna in colunas_necessarias:
            if coluna not in dados_filtrados.columns:
                dados_filtrados[coluna] = ""

        ### garantir que todas as colunas tenham valores válidos ###
        dados_filtrados['OBSERVAÇÃO'] = dados_filtrados['OBSERVAÇÃO'].fillna("")

        ### selecionar as colunas para exportação ###
        dados_exportar = dados_filtrados[colunas_necessarias + ['ESTORNOS', 'PENDENCIAS']]

        ### gerar nome do arquivo Excel baseado na data e hora atuais ###
        data_atual = datetime.now().strftime("%m%Y_%H%M")
        caminho_excel = f"Pendencias_{data_atual}.xlsx"

        ### exportar os dados filtrados para o novo arquivo Excel ###
        dados_exportar.to_excel(caminho_excel, index=False)

        return caminho_excel

    except Exception as e:
        print(f"Erro ao exportar dados de pendências: {e}")
        return None

### função principal para exibir a aplicação ###
def main(page: ft.Page):
    
    exportar_impostos_para_json()
    inserir_dados_no_excel()

    ### configurações da página ###
    page.title = "Sistema de Provisão e Estorno"
    page.horizontal_alignment = ft.CrossAxisAlignment.START
    page.vertical_alignment = ft.MainAxisAlignment.START
    page.padding = 0
    page.spacing = 0
    page.window.width = 1400
    page.window.height = 740
    page.window.center()

    ### container para o conteúdo principal ###
    conteudo_principal = ft.Container(
        content=None,  ### inicialmente vazio, será preenchido conforme as seleções do menu lateral ###
        expand=True,
        padding=0,
        bgcolor="#FFFFFF",  ### fundo branco para a área de conteúdo principal ###
        border_radius=ft.border_radius.all(0),
    )

    ### função para realizar o download das pendências ###
    def realizar_download_pendencias(e):
        caminho_arquivo = baixar_pendencias()
        if caminho_arquivo:
            ### exibe uma mensagem de sucesso com o nome do arquivo baixado ###
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Arquivo exportado para: {caminho_arquivo}")))
        else:
            ### exibe uma mensagem de erro caso algo dê errado ###
            page.show_snack_bar(ft.SnackBar(content=ft.Text(f"Erro ao exportar o arquivo."), bgcolor=ft.colors.RED))

    ### definir o menu lateral com os botões ###
    menu_lateral = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text(
                    "Gerenciamento de Provisões", 
                    color=ft.colors.WHITE,  ### texto branco acima dos botões ###
                    size=28,
                    text_align=ft.TextAlign.CENTER,
                    max_lines=2,  ### permite quebra de linha ###
                ),
                ft.Container(
                    height=30,  ### espaçamento entre o título e os botões ###
                ),
                botao_menu_lateral("Provisão", lambda _: TelaProvisao(page, conteudo_principal)),
                botao_menu_lateral("Estorno", lambda _: TelaEstorno(page, conteudo_principal)),
                botao_menu_lateral("Cliente", lambda _: TelaCliente(page, conteudo_principal)),
                botao_menu_lateral("Pendências", realizar_download_pendencias),  ### botão de Pendências estilizado como os demais ###
                ft.Container(
                    content=ft.Image(
                        src="imagens/logo.png",  ### imagem exibida no fundo do menu lateral ###
                        width=150,
                        height=150,
                        fit=ft.ImageFit.CONTAIN,
                    ),
                    alignment=ft.alignment.center,  ### centraliza a imagem ###
                    padding=ft.padding.only(top=130),  ### espaçamento da imagem em relação ao topo ###
                ),
            ],
            alignment=ft.MainAxisAlignment.START,
            spacing=10,  ### espaçamento entre os botões do menu lateral ###
            expand=True,  ### expande o menu lateral ###
        ),
        width=240,
        padding=10,
        bgcolor="#212A57",  ### cor de fundo do menu lateral ###
    )

    ### iniciar a tela de Provisão ao carregar a página ###
    TelaProvisao(page, conteudo_principal)

    ### estrutura da página com o menu lateral e o conteúdo principal ###
    page.add(
        ft.Row(
            controls=[
                menu_lateral,
                conteudo_principal,  ### o conteúdo principal será exibido ao lado do menu lateral ###
            ],
            expand=True,
            spacing=0,
        )
    )

### inicia a aplicação ###
ft.app(target=main)
